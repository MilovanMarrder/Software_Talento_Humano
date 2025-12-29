from models.kardex_dao import KardexDAO
from logics.vacation_service import VacationService
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ReportService:
    def __init__(self):
        self.kardex_dao = KardexDAO()
        self.vac_service = VacationService()

    def get_kardex_report_data(self, id_contrato, fecha_ini=None, fecha_fin=None):
        """
        Genera la estructura de datos completa para el reporte de Kardex.
        Agnóstico de la UI (sirve para Tkinter, Excel, PDF, JSON API).
        """
        response = {
            "saldo_anterior": 0.0,
            "movimientos": [], # Lista de dicts
            "totales": {"debe": 0.0, "haber": 0.0, "saldo_final": 0.0}
        }

        # 1. Procesar devengos automáticos hasta hoy (Write to BD)
        try:
            self.vac_service.process_monthly_accruals(id_contrato)
        except Exception as e:
            print(f"Error procesando devengos automáticos: {e}")

        # 2. Obtener datos históricos (Read from BD)
        saldo_ant, raw_movs = self.kardex_dao.get_kardex_report(id_contrato, fecha_ini, fecha_fin)
        response["saldo_anterior"] = saldo_ant
        
        current_balance = saldo_ant
        total_debe = 0.0
        total_haber = 0.0

        # 3. Procesar filas históricas
        # raw_movs estructura: (id, fecha, tipo, obs, dias, f_ini_real, f_fin_real)
        for m in raw_movs:
            dias = m[4]
            current_balance += dias
            
            # Formatear detalle
            detalle = m[3]
            if m[5] and m[6]:
                detalle = f"{detalle} [Del {m[5]} al {m[6]}]"

            # Clasificar Debe/Haber
            debe = abs(dias) if dias < 0 else 0.0
            haber = dias if dias > 0 else 0.0
            
            total_debe += debe
            total_haber += haber

            response["movimientos"].append({
                "fecha": m[1],
                "tipo": m[2],
                "detalle": detalle,
                "debe": debe,
                "haber": haber,
                "saldo": current_balance,
                "es_proyeccion": False
            })

        # 4. Procesar Proyecciones (Si hay fecha fin futura)
        if fecha_fin:
            proyecciones = self.vac_service.get_future_projections(id_contrato, fecha_fin)
            for p in proyecciones:
                dias = p['dias']
                current_balance += dias
                total_haber += dias # Proyección siempre suma (es ganancia futura)
                
                response["movimientos"].append({
                    "fecha": p['fecha'],
                    "tipo": "PROYECCION",
                    "detalle": p['detalle'],
                    "debe": 0.0,
                    "haber": dias,
                    "saldo": current_balance,
                    "es_proyeccion": True
                })

        # 5. Totales Finales
        response["totales"]["debe"] = total_debe
        response["totales"]["haber"] = total_haber
        response["totales"]["saldo_final"] = current_balance
        
        return response
    
    def export_kardex_excel(self, id_contrato, f_ini, f_fin, filepath, employee_name="Kardex"):
        """
        Genera un archivo Excel con el reporte de Kardex idéntico al de la pantalla.
        """
        try:
            # 1. REUTILIZAR LA LÓGICA DE CÁLCULO
            # Usamos el mismo método que alimenta la vista para asegurar consistencia
            data = self.get_kardex_report_data(id_contrato, f_ini, f_fin)
            
            # 2. CREAR LIBRO Y HOJA
            wb = openpyxl.Workbook()
            ws = wb.active
            # ws.title = "Kardex Vacaciones"

                        # --- LIMPIEZA DEL NOMBRE DE LA HOJA ---
            # Quitamos caracteres prohibidos por Excel: [ ] : * ? / \
            invalid_chars = ['[', ']', ':', '*', '?', '/', '\\']
            clean_name = employee_name
            for char in invalid_chars:
                clean_name = clean_name.replace(char, '')
            
            # Excel limita el nombre de la hoja a 31 caracteres
            ws.title = clean_name[:30] 
            
            # 3. ESTILOS
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            bold_font = Font(bold=True)
            
            # 4. ENCABEZADOS
            headers = ["Fecha", "Tipo Movimiento", "Detalle / Observación", "Debe (Devengado)", "Haber (Ganado)", "Saldo"]
            ws.append(headers)
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill

            # 5. FILA SALDO ANTERIOR (Si aplica)
            if data["saldo_anterior"] != 0 or f_ini:
                ws.append([
                    f_ini if f_ini else "---",
                    "SALDO ANTERIOR",
                    "Arrastre de periodo previo",
                    "",
                    "",
                    data['saldo_anterior']
                ])
                # Poner en negrita la fila del saldo anterior
                for cell in ws[ws.max_row]:
                    cell.font = bold_font

            # 6. MOVIMIENTOS
            for row in data["movimientos"]:
                ws.append([
                    row['fecha'],
                    row['tipo'],
                    row['detalle'],
                    row['debe'] if row['debe'] > 0 else "",
                    row['haber'] if row['haber'] > 0 else "",
                    row['saldo']
                ])
                
                # Si es proyección, quizás ponerlo en cursiva (opcional)
                if row.get('es_proyeccion'):
                    for cell in ws[ws.max_row]:
                        cell.font = Font(italic=True, color="555555")

            # 7. TOTALES
            tot = data["totales"]
            ws.append([
                "", 
                "TOTALES", 
                "",
                tot['debe'],
                tot['haber'],
                tot['saldo_final']
            ])
            # Estilo fila totales
            last_row = ws.max_row
            for col in range(1, 7):
                cell = ws.cell(row=last_row, column=col)
                cell.font = bold_font
                cell.fill = PatternFill(start_color="DCE6F1", fill_type="solid")

            # 8. AJUSTAR ANCHO DE COLUMNAS
            column_widths = [12, 25, 40, 15, 15, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # 9. GUARDAR
            wb.save(filepath)
            return True, "Reporte exportado correctamente."

        except Exception as e:
            return False, f"Error al exportar Excel: {str(e)}"